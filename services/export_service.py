from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from services.config import EXCEL_PATH
from services.storage_service import read_saved_numbers


def export_saved_numbers_to_excel() -> Path:
    rows = read_saved_numbers()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Numéros extraits"

    visible_columns = ["date", "fichier", "numero", "type", "confiance"]
    headers = ["Date", "Fichier", "Numéro", "Type détecté", "Confiance"]
    worksheet.append(headers)

    header_fill = PatternFill("solid", fgColor="168A52")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        worksheet.append([row.get(column, "") for column in visible_columns])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    if worksheet.max_row >= 2:
        confidence_column = "E"
        confidence_range = f"{confidence_column}2:{confidence_column}{worksheet.max_row}"
        worksheet.conditional_formatting.add(
            confidence_range,
            CellIsRule(operator="greaterThanOrEqual", formula=["85"], fill=PatternFill("solid", fgColor="DCFCE7")),
        )
        worksheet.conditional_formatting.add(
            confidence_range,
            CellIsRule(operator="between", formula=["65", "84"], fill=PatternFill("solid", fgColor="FEF3C7")),
        )
        worksheet.conditional_formatting.add(
            confidence_range,
            CellIsRule(operator="lessThan", formula=["65"], fill=PatternFill("solid", fgColor="FEE2E2")),
        )

    for row_cells in worksheet.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="center")

    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        column_letter = get_column_letter(column_cells[0].column)
        worksheet.column_dimensions[column_letter].width = min(max_length + 4, 46)

    export_path = EXCEL_PATH.with_name(f"screenpict_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    workbook.save(export_path)
    return export_path
